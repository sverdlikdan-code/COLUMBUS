# -*- coding: utf-8 -*-
"""
SADRAN analysis builder.
Reads C:\\Users\\d.sverdlik\\Desktop\\SADRAN.xlsx (Export sheet, already merged with כשרות/עיר by user)
and writes a NEW workbook SADRAN_ANALYSIS.xlsx with growth/decline analysis by
department, merchandiser (sadran), customer and kosher-market split.
Original SADRAN.xlsx is left untouched (it contains a live PivotTable in Sheet1).

Чтение источника — openpyxl (read_only). Запись результата — xlsxwriter, НЕ openpyxl:
на масштабе этой книги (12+ Excel Table на разных листах, тысячи строк) openpyxl
надёжно портит styles.xml/tables ("We found a problem with some content..." в Excel),
а xlsxwriter на той же схеме и объёме данных открывается чисто (проверено эмпирически
на синтетическом тесте той же структуры) — см. excel-smart-reports skill.
"""
import re
import json
import openpyxl
import xlsxwriter
from collections import defaultdict as _dd

# fixBiDi — порт из server/index.js:2110 (hebrew-bidi skill).
# Имена клиентов из PBI приходят обёрнутые в LRO/PDF-маркеры вокруг REVERSE()'нутой строки
# (см. reference/sql-customers-form.sql: NCHAR(8237)+REVERSE(CUSTDES)+NCHAR(8236)) —
# без этого фикса имя отображается задом наперёд.
_BIDI_TEST = re.compile(r"[\u200E\u200F\u202A-\u202E]")
_BIDI_STRIP = re.compile(r"[\u200E\u200F\u202A-\u202E]")
_HEB = re.compile(r"[א-ת]")
_DIGITS = re.compile(r"\d+")

def fix_bidi(raw):
    if not raw:
        return ""
    has_bidi = bool(_BIDI_TEST.search(raw))
    s = _BIDI_STRIP.sub("", raw).strip()
    if not has_bidi or not _HEB.search(s):
        return s
    words = s.split()
    words.reverse()
    fixed_words = []
    for w in words:
        if _HEB.search(w):
            w2 = w[::-1]
            w2 = _DIGITS.sub(lambda m: m.group()[::-1], w2)
            fixed_words.append(w2)
        else:
            fixed_words.append(w)
    fixed = " ".join(fixed_words)
    fixed = fixed.replace("(", "\x01").replace(")", "(").replace("\x01", ")")
    return fixed

SRC = r"C:\Users\d.sverdlik\Desktop\SADRAN.xlsx"
OUT = r"C:\Users\d.sverdlik\Desktop\SADRAN_ANALYSIS.xlsx"

# Без defusedxml: SRC — файл самого пользователя с его Desktop, не вложение из письма/скачанное
# (см. excel-smart-reports skill, Правило №7 — defusedxml нужен для НЕДОВЕРЕННЫХ источников).
wb_src = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws_src = wb_src["Export"]

# מחלקה -> компания (утверждено пользователем 2026-07-21, рукописная разметка):
# mish גלידה -> ICE MISH, גלידה bdd -> ICE BDD, מדף/מתוקים -> INTER,
# דג יבש/דגים/חלבי/קפוא -> FORMULA
DEPT_COMPANY = {
    "mish גלידה": "ICE MISH",
    "גלידה bdd": "ICE BDD",
    "מדף": "INTER",
    "מתוקים": "INTER",
    "דג יבש": "FORMULA",
    "דגים": "FORMULA",
    "חלבי": "FORMULA",
    "קפוא ❄": "FORMULA",
}

# סוג לקוח (тип клиента) — из FORMULA PBI dataset, таблица "לקוחות FORM+I+INT", join по מס. לקוח
with open(r"C:\Users\d.sverdlik\Desktop\WORKSPACE\COLUMBUS\scripts\cust_type_dump.json", encoding="utf-8") as f:
    _cust_type_raw = json.load(f)
CUST_TYPE = {str(r["[custno]"]).strip(): (r["[custtype]"] or "(не указано)") for r in _cust_type_raw}

rows = []
for r in ws_src.iter_rows(min_row=3, values_only=True):
    kosher, city, custno, custname, sadran, dept, last_year, now, pct = r[1:10]
    if custno is None:
        continue
    dept_clean = (dept or "").strip()
    rows.append({
        "kosher": kosher, "city": city, "custno": custno, "custname": fix_bidi(custname),
        "sadran": sadran, "dept": dept_clean,
        "company": DEPT_COMPANY.get(dept_clean, "(не определено)"),
        "custtype": CUST_TYPE.get(str(custno).strip(), "(нет в PBI)"),
        "last_year": last_year or 0, "now": now or 0,
    })

def pct_change(ly, now):
    if ly:
        return (now - ly) / ly
    return None

def agg(rows, keyfunc):
    buckets = {}
    for row in rows:
        k = keyfunc(row)
        b = buckets.setdefault(k, {"last_year": 0.0, "now": 0.0, "n": 0})
        b["last_year"] += row["last_year"]
        b["now"] += row["now"]
        b["n"] += 1
    out = []
    for k, v in buckets.items():
        out.append({
            "key": k, "last_year": v["last_year"], "now": v["now"],
            "delta": v["now"] - v["last_year"],
            "pct": pct_change(v["last_year"], v["now"]),
            "n": v["n"],
        })
    return out

by_company = sorted(agg(rows, lambda r: r["company"]), key=lambda x: x["delta"], reverse=True)
by_dept_raw = agg(rows, lambda r: (r["company"], r["dept"]))
by_dept = sorted(by_dept_raw, key=lambda x: (x["key"][0], -x["delta"]))

# Same-Store классификация клиентов (глобально, по всем их строкам) — принято в FMCG-дистрибуции:
# % роста по סדרן считать только по клиентам с историей (like-for-like), иначе новый клиент
# искусственно раздувает "рост". Новых клиентов сдаранам НАЗНАЧАЮТ (территория/канал), они их
# не приводят сами — смешивать эту сумму с ростом на своей базе в один % некорректно вдвойне.
_cust_ly_total = {}
for _r in rows:
    _cust_ly_total[_r["custno"]] = _cust_ly_total.get(_r["custno"], 0.0) + _r["last_year"]
NEW_CUSTOMERS_GLOBAL = {c for c, ly in _cust_ly_total.items() if ly == 0}

# ICE BDD исключаем из ВСЕХ разрезов по сдаранам и по клиентам (утверждено пользователем
# 2026-07-21): этот товар идёт через канал OneSales (водитель=агент=мерч), сдараны его не
# ведут — включать его в их персональные показатели и в клиентские топы некорректно.
# Компания-уровень (Summary/По компаниям/Same-Store vs Новые) — ICE BDD там остаётся,
# это реальная выручка, просто не относится к работе сдаранов.
rows_ex_bdd = [r for r in rows if r["company"] != "ICE BDD"]

same_store_rows = [r for r in rows_ex_bdd if r["custno"] not in NEW_CUSTOMERS_GLOBAL]
new_cust_rows = [r for r in rows_ex_bdd if r["custno"] in NEW_CUSTOMERS_GLOBAL]

by_sadran_ss = sorted(agg(same_store_rows, lambda r: r["sadran"]), key=lambda x: x["delta"], reverse=True)
by_sadran_new = {b["key"]: b for b in agg(new_cust_rows, lambda r: r["sadran"])}
by_sadran = by_sadran_ss  # используется везде ниже как основной разрез по сдаранам (same-store, без ICE BDD)
by_kosher = sorted(agg(rows, lambda r: r["kosher"] or "(не указано)"), key=lambda x: x["delta"], reverse=True)
by_kosher_dept = sorted(agg(rows, lambda r: (r["kosher"] or "(не указано)", r["dept"])), key=lambda x: x["delta"], reverse=True)
by_sadran_dept = sorted(agg(same_store_rows, lambda r: (r["sadran"], r["dept"])), key=lambda x: x["delta"], reverse=True)
by_sadran_dept_new = {b["key"]: b["now"] for b in agg(new_cust_rows, lambda r: (r["sadran"], r["dept"]))}
by_city = sorted(agg(rows, lambda r: r["city"] or "(не указан)"), key=lambda x: x["delta"], reverse=True)
by_custtype = sorted(agg(rows, lambda r: r["custtype"]), key=lambda x: x["delta"], reverse=True)
by_custtype_company = sorted(agg(rows, lambda r: (r["custtype"], r["company"])), key=lambda x: (x["key"][0], -x["delta"]))

# Клиентские топы/риски — тоже без ICE BDD (см. выше)
by_customer = agg(rows_ex_bdd, lambda r: (r["custno"], r["custname"], r["sadran"], r["kosher"], r["city"]))
by_customer_sorted = sorted(by_customer, key=lambda x: x["delta"], reverse=True)
top_growth = by_customer_sorted[:25]
top_decline = list(reversed(by_customer_sorted[-25:]))

# ---------- FORMULA: топ-10 клиентов + динамика по департаментам внутри каждого ----------
formula_rows = [r for r in rows if r["company"] == "FORMULA"]
formula_cust_totals = _dd(lambda: {"ly": 0.0, "now": 0.0, "custname": None})
for r in formula_rows:
    c = formula_cust_totals[r["custno"]]
    c["ly"] += r["last_year"]; c["now"] += r["now"]; c["custname"] = r["custname"]
formula_top10 = sorted(formula_cust_totals.items(), key=lambda x: x[1]["now"], reverse=True)[:10]
formula_dept_by_cust = _dd(lambda: _dd(lambda: {"ly": 0.0, "now": 0.0}))
for r in formula_rows:
    d = formula_dept_by_cust[r["custno"]][r["dept"]]
    d["ly"] += r["last_year"]; d["now"] += r["now"]

# ================== запись книги (xlsxwriter) ==================
wb = xlsxwriter.Workbook(OUT, {"strings_to_formulas": False})

NAVY = "#1F4E78"
GROWTH_BG, GROWTH_FG = "#C6EFCE", "#006100"
DECLINE_BG, DECLINE_FG = "#FFC7CE", "#9C0006"

header_fmt = wb.add_format({"bold": True, "bg_color": NAVY, "font_color": "white", "align": "center", "valign": "vcenter"})
bold_fmt = wb.add_format({"bold": True})
title_fmt = wb.add_format({"bold": True, "font_size": 14})
subtitle_fmt = wb.add_format({"bold": True, "font_size": 13, "font_color": NAVY})
money_fmt = wb.add_format({"num_format": "#,##0"})
money_bold_fmt = wb.add_format({"num_format": "#,##0", "bold": True})
pct_fmt = wb.add_format({"num_format": "0.0%"})
pct_growth_fmt = wb.add_format({"num_format": "0.0%", "bg_color": GROWTH_BG, "font_color": GROWTH_FG})
pct_decline_fmt = wb.add_format({"num_format": "0.0%", "bg_color": DECLINE_BG, "font_color": DECLINE_FG})
pct_growth_bold_fmt = wb.add_format({"num_format": "0.0%", "font_color": GROWTH_FG, "bold": True})
pct_decline_bold_fmt = wb.add_format({"num_format": "0.0%", "font_color": DECLINE_FG, "bold": True})
note_fmt = wb.add_format({"italic": True, "font_size": 10, "font_color": "#808080", "text_wrap": True})
reco_header_fmt = wb.add_format({"bold": True, "font_size": 13, "font_color": NAVY, "text_wrap": True, "valign": "top"})
reco_body_fmt = wb.add_format({"font_size": 11, "text_wrap": True, "valign": "top"})

def pct_val_fmt(val):
    if val is None:
        return pct_fmt
    return pct_growth_fmt if val > 0 else (pct_decline_fmt if val < 0 else pct_fmt)

def write_table(ws, start_row0, headers, data_rows, pct_col_idx=None, money_cols=None, table_name=None):
    """start_row0 = 0-indexed header row. pct_col_idx/money_cols — 1-indexed column numbers
    (как в исходной openpyxl-версии), для совместимости с остальным кодом."""
    ncols = len(headers)
    money_cols = set(money_cols or [])
    r = start_row0 + 1
    for row in data_rows:
        for i, val in enumerate(row):
            col_1based = i + 1
            if val is None:
                continue
            if pct_col_idx and col_1based == pct_col_idx and isinstance(val, (int, float)):
                ws.write_number(r, i, val, pct_val_fmt(val))
            elif col_1based in money_cols and isinstance(val, (int, float)):
                ws.write_number(r, i, val, money_fmt)
            else:
                # Formula-injection защита — на уровне Workbook({"strings_to_formulas": False})
                # ниже, НЕ через write_string() (ломает файл на масштабе >10 Table в xlsxwriter
                # 3.2.9 — подтверждено эмпирически 2026-07-22, см. excel-smart-reports skill).
                ws.write(r, i, val)
        r += 1
    last_row = r - 1
    if data_rows:
        ws.add_table(start_row0, 0, last_row, ncols - 1, {
            "columns": [{"header": h} for h in headers],
            "style": "Table Style Medium 2",
            "autofilter": True,
            "name": table_name,
        })
    for i, h in enumerate(headers):
        maxlen = max([len(str(h))] + [len(str(row[i])) for row in data_rows if row[i] is not None]) if data_rows else len(str(h))
        ws.set_column(i, i, min(max(maxlen + 2, 10), 45))
    ws.freeze_panes(start_row0 + 1, 0)
    return last_row + 1  # следующая свободная 0-indexed строка

_tbl_seq = [0]
def next_table_name():
    _tbl_seq[0] += 1
    return f"Tbl{_tbl_seq[0]}"

def add_bar_chart(ws, title, sheet_name, cat_col0, val_col0, first_row0, last_row0, anchor="H1"):
    chart = wb.add_chart({"type": "bar"})
    chart.add_series({
        "name": title,
        "categories": [sheet_name, first_row0, cat_col0, last_row0, cat_col0],
        "values": [sheet_name, first_row0, val_col0, last_row0, val_col0],
    })
    chart.set_title({"name": title})
    chart.set_y_axis({"name": "₪"})
    chart.set_legend({"none": True})
    ws.insert_chart(anchor, chart, {"x_scale": 1.6, "y_scale": 1.6})

# ---------- Sheet: Summary ----------
ws = wb.add_worksheet("Summary")
ws.right_to_left()
total_ly = sum(r["last_year"] for r in rows)
total_now = sum(r["now"] for r in rows)
total_pct = pct_change(total_ly, total_now)
ws.write(1, 1, "SADRAN — анализ роста/падения продаж (2026 vs 2025)", title_fmt)
ws.write(3, 1, "Всего продаж, прошлый период", bold_fmt)
ws.write_number(3, 2, total_ly, money_fmt)
ws.write(4, 1, "Всего продаж, текущий период", bold_fmt)
ws.write_number(4, 2, total_now, money_fmt)
ws.write(5, 1, "Изменение", bold_fmt)
ws.write_number(5, 2, total_pct, pct_val_fmt(total_pct))
ws.set_column(1, 1, 32)
ws.set_column(2, 2, 16)

next_row = write_table(
    ws, 8,
    ["Кошерность", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_kosher],
    pct_col_idx=5, money_cols={2, 3, 4}, table_name=next_table_name(),
)
add_bar_chart(ws, "Продажи по кошерности — Δ", "Summary", 0, 3, 8, next_row - 1, anchor="H9")

# ---------- Sheet: По компаниям (חברה) — верхний уровень ----------
ws_co = wb.add_worksheet("По компаниям")
ws_co.right_to_left()
next_row_co = write_table(
    ws_co, 0,
    ["Компания (חברה)", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк",
     "Доля в total (прошлый)", "Доля в total (текущий)"],
    [[b["key"], b["last_year"], b["now"], b["delta"], b["pct"], b["n"],
      (b["last_year"] / total_ly) if total_ly else None,
      (b["now"] / total_now) if total_now else None] for b in by_company],
    pct_col_idx=5, money_cols={2, 3, 4}, table_name=next_table_name(),
)
for i, b in enumerate(by_company):
    rr_ = 1 + i
    if total_ly:
        ws_co.write_number(rr_, 6, b["last_year"] / total_ly, pct_fmt)
    if total_now:
        ws_co.write_number(rr_, 7, b["now"] / total_now, pct_fmt)
add_bar_chart(ws_co, "Δ продаж по компаниям", "По компаниям", 0, 3, 0, next_row_co - 1, anchor="J1")

# ---------- Sheet: По департаментам (внутри каждой компании) ----------
ws2 = wb.add_worksheet("По департаментам")
ws2.right_to_left()
next_row2 = write_table(
    ws2, 0,
    ["Компания (חברה)", "Департамент (מחלקה)", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"][0], b["key"][1], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_dept],
    pct_col_idx=6, money_cols={3, 4, 5}, table_name=next_table_name(),
)
add_bar_chart(ws2, "Δ продаж по департаментам", "По департаментам", 1, 4, 0, next_row2 - 1, anchor="H1")

# ---------- Sheet: По кошерности x департамент ----------
ws3 = wb.add_worksheet("Кошерность x Департамент")
ws3.right_to_left()
write_table(
    ws3, 0,
    ["Кошерность", "Департамент", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"][0], b["key"][1], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_kosher_dept],
    pct_col_idx=6, money_cols={3, 4, 5}, table_name=next_table_name(),
)

# ---------- Sheet: По сдаранам (Same-Store — принято в FMCG-дистрибуции) ----------
ws4 = wb.add_worksheet("По сдаранам")
ws4.right_to_left()
ws4.write(0, 7, ("% изменение считается ТОЛЬКО по клиентам с историей прошлого года (like-for-like/"
                  "same-store) — новые клиенты не искажают % роста, их вклад показан отдельной колонкой "
                  "(₪, без истории для сравнения)."), note_fmt)
ws4.set_column(7, 7, 55)

def _new_note(sadran_key):
    b = by_sadran_new.get(sadran_key)
    if not b or not b["now"]:
        return None
    return b["now"]

next_row4 = write_table(
    ws4, 0,
    ["Сдаран (סדרן)", "Прошлый период (Same-Store)", "Текущий период (Same-Store)", "Δ (сумма)",
     "% изменение (Same-Store)", "# строк", "+ Новые клиенты (₪, без истории)"],
    [[b["key"], b["last_year"], b["now"], b["delta"], b["pct"], b["n"], _new_note(b["key"])] for b in by_sadran],
    pct_col_idx=5, money_cols={2, 3, 4, 7}, table_name=next_table_name(),
)
add_bar_chart(ws4, "Δ продаж по сдаранам", "По сдаранам", 0, 3, 0, next_row4 - 1, anchor="I1")

# ---------- Sheet: Сдаран x Департамент (Same-Store) ----------
ws5 = wb.add_worksheet("Сдаран x Департамент")
ws5.right_to_left()
ws5.write(0, 8, ("% изменение — только клиенты с историей прошлого года (same-store/like-for-like). "
                  "Новые клиенты — отдельной колонкой, вне %."), note_fmt)
ws5.set_column(8, 8, 55)
write_table(
    ws5, 0,
    ["Сдаран", "Департамент", "Прошлый период (SS)", "Текущий период (SS)", "Δ (сумма)",
     "% изменение (SS)", "# строк", "+ Новые клиенты (₪)"],
    [[b["key"][0], b["key"][1], b["last_year"], b["now"], b["delta"], b["pct"], b["n"],
      by_sadran_dept_new.get(b["key"])] for b in by_sadran_dept],
    pct_col_idx=6, money_cols={3, 4, 5, 8}, table_name=next_table_name(),
)

# ---------- Sheet: По городам (רынки) ----------
ws_city = wb.add_worksheet("По городам")
ws_city.right_to_left()
write_table(
    ws_city, 0,
    ["Город (עיר)", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_city],
    pct_col_idx=5, money_cols={2, 3, 4}, table_name=next_table_name(),
)

# ---------- Sheet: По типу клиента (סוג לקוח, из FORMULA PBI) ----------
ws_type = wb.add_worksheet("По типу клиента")
ws_type.right_to_left()
next_row_type = write_table(
    ws_type, 0,
    ["Тип клиента (סוג לקוח)", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_custtype],
    pct_col_idx=5, money_cols={2, 3, 4}, table_name=next_table_name(),
)
add_bar_chart(ws_type, "Δ продаж по типу клиента", "По типу клиента", 0, 3, 0, next_row_type - 1, anchor="H1")

# ---------- Sheet: Тип клиента x Компания ----------
ws_type_co = wb.add_worksheet("Тип клиента x Компания")
ws_type_co.right_to_left()
write_table(
    ws_type_co, 0,
    ["Тип клиента", "Компания", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение", "# строк"],
    [[b["key"][0], b["key"][1], b["last_year"], b["now"], b["delta"], b["pct"], b["n"]] for b in by_custtype_company],
    pct_col_idx=6, money_cols={3, 4, 5}, table_name=next_table_name(),
)

# ---------- Sheet: Клиенты — рост / падение ----------
ws6 = wb.add_worksheet("Клиенты — рост")
ws6.right_to_left()
write_table(
    ws6, 0,
    ["№ клиента", "Клиент", "Сдаран", "Кошерность", "Город", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение"],
    [[b["key"][0], b["key"][1], b["key"][2], b["key"][3], b["key"][4], b["last_year"], b["now"], b["delta"], b["pct"]] for b in top_growth],
    pct_col_idx=9, money_cols={6, 7, 8}, table_name=next_table_name(),
)

ws7 = wb.add_worksheet("Клиенты — падение")
ws7.right_to_left()
write_table(
    ws7, 0,
    ["№ клиента", "Клиент", "Сдаран", "Кошерность", "Город", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение"],
    [[b["key"][0], b["key"][1], b["key"][2], b["key"][3], b["key"][4], b["last_year"], b["now"], b["delta"], b["pct"]] for b in top_decline],
    pct_col_idx=9, money_cols={6, 7, 8}, table_name=next_table_name(),
)

# ---------- Разбивка: текучка / новые / рост-падение существующих (по компаниям) ----------
churn_buckets = {}
for r in rows:
    b = churn_buckets.setdefault(r["company"], {
        "churn": 0.0, "churn_n": 0, "new": 0.0, "new_n": 0,
        "decline": 0.0, "decline_n": 0, "growth": 0.0, "growth_n": 0,
    })
    ly, now = r["last_year"], r["now"]
    if ly > 0 and now == 0:
        b["churn"] += ly; b["churn_n"] += 1
    elif ly == 0 and now > 0:
        b["new"] += now; b["new_n"] += 1
    elif ly > 0 and now > 0 and now < ly:
        b["decline"] += (now - ly); b["decline_n"] += 1
    elif ly > 0 and now > 0 and now > ly:
        b["growth"] += (now - ly); b["growth_n"] += 1

ws8 = wb.add_worksheet("Текучка-Новые-Существующие")
ws8.right_to_left()
churn_rows = []
for co in ["FORMULA", "INTER", "ICE MISH", "ICE BDD"]:
    b = churn_buckets.get(co)
    if not b:
        continue
    churn_rows.append([
        co, -b["churn"], b["churn_n"], b["new"], b["new_n"],
        b["decline"], b["decline_n"], b["growth"], b["growth_n"],
    ])
write_table(
    ws8, 0,
    ["Компания", "Потеряно (ушедшие клиенты)", "# ушедших", "Новые клиенты (₪)", "# новых",
     "Существующие: падение (₪)", "# упавших", "Существующие: рост (₪)", "# выросших"],
    churn_rows, money_cols={2, 4, 6, 8}, table_name=next_table_name(),
)

# ---------- Отток — детали (поимённо, кто и по какому מחלקה ушёл в 0) ----------
ws8b = wb.add_worksheet("Отток — детали")
ws8b.right_to_left()
churn_detail_rows = []
for r in rows:
    if r["last_year"] > 0 and r["now"] == 0:
        churn_detail_rows.append([
            r["custno"], r["custname"], r["dept"], r["company"], r["sadran"], r["city"], r["last_year"],
        ])
churn_detail_rows.sort(key=lambda row: -row[6])
write_table(
    ws8b, 0,
    ["№ клиента", "Клиент", "Департамент", "Компания", "Сдаран", "Город", "Было (₪, стало 0)"],
    churn_detail_rows, money_cols={7}, table_name=next_table_name(),
)

# ---------- Same-Store vs Новые клиенты ----------
cust_totals = _dd(lambda: {"ly": 0.0, "now": 0.0, "custname": None})
for r in rows:
    c = cust_totals[r["custno"]]
    c["ly"] += r["last_year"]; c["now"] += r["now"]; c["custname"] = r["custname"]

new_customers = {c for c, v in cust_totals.items() if v["ly"] == 0 and v["now"] > 0}
same_store_customers = {c for c, v in cust_totals.items() if v["ly"] > 0}

same_ly_total = sum(cust_totals[c]["ly"] for c in same_store_customers)
same_now_total = sum(cust_totals[c]["now"] for c in same_store_customers)
new_now_total = sum(cust_totals[c]["now"] for c in new_customers)
grand_now = sum(v["now"] for v in cust_totals.values())

ws10 = wb.add_worksheet("Same-Store vs Новые клиенты")
ws10.right_to_left()
ws10.write(1, 1, "Растём ли только за счёт новых клиентов/рынков, или органически?", subtitle_fmt)
ws10.set_column(1, 1, 30)
ws10.set_column(2, 2, 16)
overview_rows = [
    ("Прошлый период (Same-Store)", same_ly_total, False),
    ("Текущий период (Same-Store)", same_now_total, False),
    ("Δ Same-Store (₪)", same_now_total - same_ly_total, False),
    ("% изменение Same-Store", pct_change(same_ly_total, same_now_total) if same_ly_total else None, True),
    ("Новые клиенты — вклад в текущий период (₪)", new_now_total, False),
    ("Доля новых клиентов в общей выручке", (new_now_total / grand_now) if grand_now else None, True),
]
rr0 = 3
for label, val, is_pct in overview_rows:
    ws10.write(rr0, 1, label, bold_fmt)
    if val is not None:
        ws10.write_number(rr0, 2, val, pct_val_fmt(val) if is_pct else money_fmt)
    rr0 += 1

company_ss_rows = []
for co in ["FORMULA", "INTER", "ICE MISH", "ICE BDD"]:
    co_rows = [x for x in rows if x["company"] == co]
    co_new_now = sum(x["now"] for x in co_rows if x["custno"] in new_customers)
    co_same_ly = sum(x["last_year"] for x in co_rows if x["custno"] in same_store_customers)
    co_same_now = sum(x["now"] for x in co_rows if x["custno"] in same_store_customers)
    co_total_now = sum(x["now"] for x in co_rows)
    pct = pct_change(co_same_ly, co_same_now)
    company_ss_rows.append([
        co, co_same_ly, co_same_now, pct, co_new_now,
        (co_new_now / co_total_now) if co_total_now else None,
    ])
write_table(
    ws10, rr0 + 1,
    ["Компания", "Same-Store: прошлый период", "Same-Store: текущий", "Same-Store % изменение",
     "Новые клиенты (₪)", "Доля новых клиентов в компании"],
    company_ss_rows, pct_col_idx=4, money_cols={2, 3, 5}, table_name=next_table_name(),
)
# доля новых клиентов (колонка 6) тоже процент — форматируем отдельно, т.к. write_table не знает про неё
for i, row in enumerate(company_ss_rows):
    val = row[5]
    if val is not None:
        ws10.write_number(rr0 + 2 + i, 5, val, pct_fmt)

# ---------- Pareto 70% — риски у ключевых клиентов ----------
ranked_customers = sorted(cust_totals.items(), key=lambda x: x[1]["now"], reverse=True)
cum = 0.0
pareto_set = set()
for custno, v in ranked_customers:
    cum += v["now"]
    pareto_set.add(custno)
    if cum >= 0.70 * grand_now:
        break

dept_by_cust = _dd(lambda: _dd(lambda: {"ly": 0.0, "now": 0.0}))
for r_ in rows:
    if r_["custno"] in pareto_set:
        d = dept_by_cust[r_["custno"]][r_["dept"]]
        d["ly"] += r_["last_year"]; d["now"] += r_["now"]

pareto_flags = []
for custno in pareto_set:
    v = cust_totals[custno]
    overall_pct = pct_change(v["ly"], v["now"])
    problems = []
    for dept, dv in dept_by_cust[custno].items():
        if dv["ly"] > 0 and dv["now"] == 0:
            problems.append(f"{dept}: остановился (было {round(dv['ly']):,})".replace(",", " "))
        elif dv["ly"] > 0 and dv["now"] > 0 and (dv["now"] - dv["ly"]) / dv["ly"] < -0.30:
            problems.append(f"{dept}: {round((dv['now']-dv['ly'])/dv['ly']*100)}%")
    if problems:
        pareto_flags.append({
            "custno": custno, "custname": v["custname"], "now": v["now"],
            "overall_pct": overall_pct, "problems": "; ".join(problems),
        })
pareto_flags.sort(key=lambda x: x["now"], reverse=True)

ws11 = wb.add_worksheet("Pareto 70% — риски")
ws11.right_to_left()
write_table(
    ws11, 0,
    ["№ клиента", "Клиент", "Текущие продажи (₪)", "Общее % изменение", "Провал по департаменту (>=30% или стоп)"],
    [[p["custno"], p["custname"], p["now"], p["overall_pct"], p["problems"]] for p in pareto_flags],
    pct_col_idx=4, money_cols={3}, table_name=next_table_name(),
)
ws11.write(0, 7, (f"Топ-{len(pareto_set)} клиентов формируют 70% текущей выручки. "
                   f"Из них {len(pareto_flags)} имеют провал >=30% (или полную остановку) хотя бы в одном מחלקה, "
                   f"даже если их общий итог положительный."), note_fmt)
ws11.set_column(7, 7, 60)

# ---------- Bottom-10 существующих клиентов — кандидаты на снятие мерчендайзера ----------
cust_sadran = {}
cust_city = {}
for r_ in rows:
    if r_["custno"] not in cust_sadran:
        cust_sadran[r_["custno"]] = r_["sadran"]
        cust_city[r_["custno"]] = r_["city"]

existing_ranked_low = sorted(
    (c for c in same_store_customers),
    key=lambda c: cust_totals[c]["now"],
)[:10]

ws12 = wb.add_worksheet("Bottom-10 — снять мерч.")
ws12.right_to_left()
bottom10_rows = []
for c in existing_ranked_low:
    v = cust_totals[c]
    bottom10_rows.append([
        c, v["custname"], cust_city.get(c), cust_sadran.get(c),
        v["ly"], v["now"], pct_change(v["ly"], v["now"]),
    ])
write_table(
    ws12, 0,
    ["№ клиента", "Клиент", "Город", "Сдаран", "Прошлый период", "Текущий период", "% изменение"],
    bottom10_rows, pct_col_idx=7, money_cols={5, 6}, table_name=next_table_name(),
)
ws12.write(0, 8, ("Существующие клиенты (есть история за прошлый год) с наименьшими текущими продажами — "
                   "кандидаты на то, чтобы забрать мерчендайзера с точки (низкая отдача от покрытия)."), note_fmt)
ws12.set_column(8, 8, 55)

# ---------- FORMULA — Топ-10 клиентов + динамика по департаментам внутри ----------
ws13 = wb.add_worksheet("FORMULA Топ-10")
ws13.right_to_left()
ws13.write(0, 6, ("Топ-10 клиентов FORMULA по текущим продажам, с разбивкой по каждому "
                   "департаменту (דגים/חלבי/דג יבש/קפוא) внутри. Строка 'ИТОГО' — сумма клиента "
                   "по всем департаментам FORMULA."), note_fmt)
ws13.set_column(6, 6, 55)
formula_top10_rows = []
for custno, v in formula_top10:
    formula_top10_rows.append([
        v["custname"], "ИТОГО", v["ly"], v["now"], v["now"] - v["ly"], pct_change(v["ly"], v["now"]),
    ])
    depts_here = sorted(formula_dept_by_cust[custno].items(), key=lambda x: x[1]["now"], reverse=True)
    for dept, d in depts_here:
        if d["ly"] == 0 and d["now"] == 0:
            continue
        formula_top10_rows.append([
            "", dept, d["ly"], d["now"], d["now"] - d["ly"], pct_change(d["ly"], d["now"]),
        ])
write_table(
    ws13, 0,
    ["Клиент", "Департамент", "Прошлый период", "Текущий период", "Δ (сумма)", "% изменение"],
    formula_top10_rows, pct_col_idx=6, money_cols={3, 4, 5}, table_name=next_table_name(),
)

# ---------- Рекомендации ----------
ws9 = wb.add_worksheet("Рекомендации")
ws9.right_to_left()
ws9.write(1, 1, "Выводы и рекомендации (на основе данных SADRAN.xlsx)", title_fmt)
ws9.set_column(1, 1, 3)
ws9.set_column(2, 2, 100)

_ss_pct = pct_change(same_ly_total, same_now_total)
_ss_pct_str = f"{_ss_pct*100:.1f}%" if _ss_pct is not None else "н/д"
_new_share_str = f"{(new_now_total/grand_now*100):.1f}%" if grand_now else "н/д"

# קפוא (заморозка) — маскировка волатильности городов за плоским итогом по компании
_kp_by_city = agg([r for r in rows if r["dept"] == "קפוא ❄"], lambda r: r["city"])
_kp_declines = sorted([b for b in _kp_by_city if b["last_year"] > 10000 and b["pct"] is not None and b["pct"] < -0.15],
                       key=lambda x: x["delta"])[:5]
_kp_growth = sorted([b for b in _kp_by_city if b["last_year"] > 10000 and b["pct"] is not None and b["pct"] > 0.15],
                     key=lambda x: -x["delta"])[:4]

reco_text = [
    ("Рост органический, а не только за счёт новых клиентов", True),
    (f"Новые клиенты (не было истории прошлого года) — это всего {len(new_customers)} из "
     f"{len(cust_totals)} клиентов и дают {_new_share_str} текущей выручки "
     f"(₪{new_now_total:,.0f}).".replace(",", " "), False),
    (f"Same-Store ({len(same_store_customers)} существующих клиентов, у которых была история): "
     f"₪{same_ly_total:,.0f} → ₪{same_now_total:,.0f} ({_ss_pct_str}). "
     "Даже если полностью убрать новых клиентов из расчёта, рост всё равно есть и по величине "
     f"почти совпадает с общим (+15.8% overall vs {_ss_pct_str} same-store) — значит компания "
     "растёт за счёт существующей базы, новые клиенты не главный драйвер.".replace(",", " "), False),
    ("Особенно показательно для INTER: same-store рост там +44.7% САМ ПО СЕБЕ (без учёта новых "
     "клиентов) — то есть существующие точки продаж стали покупать намного больше מדף/מתוקים, "
     "это не эффект от открытия новых точек.", False),
    ("", False),
    ("Риск концентрации: 76 клиентов (30% базы) дают 70% выручки — 19 из них уже проседают", True),
    ("В листе 'Pareto 70% — риски' — конкретные ключевые клиенты с провалом >=30% (или полной "
     "остановкой) в отдельном מחלקה, даже когда их итоговая сумма в плюсе (провал в одной категории "
     "маскируется ростом в другой).", False),
    ("Заметный сетевой паттерн: сеть 'קשת טעמים' (6 филиалов — ראשון לציון, פתח תקווה, גבעת ברנר, "
     "כפר סבא, נתניה, אשדוד) синхронно теряет именно מדף — падение от -32% до -58% в каждом филиале, "
     "при этом их общий итог у некоторых даже положительный за счёт других департаментов. Похоже на "
     "решение по этой категории на уровне сети/закупки, а не на разовую историю по одной точке — "
     "стоит выяснить у сети причину именно по מדף.", False),
    ("Второй паттерн: сеть 'קרל ברג' (филиалы אשדוד и לוד) теряет mish גלידה: -63% и -85% "
     "соответственно — тоже похоже на решение по категории на уровне сети, стоит проверить отдельно "
     "от общей картины по mish גלידה.", False),
    ("", False),
    ("קפוא выглядит стабильным (+3.6%) только в среднем по компании — по городам это иллюзия", True),
    (
        "Плюс 3.6% по FORMULA/קפוא ❄ на самом деле складывается из сильного роста в одних городах и "
        "резкого падения в других — они гасят друг друга в общем итоге. Растут: "
        + ", ".join(f"{b['key']} (+{b['pct']*100:.0f}%, +₪{b['delta']:,.0f})".replace(",", " ") for b in _kp_growth)
        + ". Падают: "
        + ", ".join(f"{b['key']} ({b['pct']*100:.0f}%, ₪{b['delta']:,.0f})".replace(",", " ") for b in _kp_declines)
        + ". Разброс в 2-3 раза больше итогового +3.6% — усреднение прячет реальную картину по קפוא.",
        False,
    ),
    ("", False),
    ("ICE BDD — единственная компания в минусе (-3.3%, -₪150K)", True),
    ("Причина НЕ в оттоке клиентов: потеряно всего ₪31K (9 клиентов). Основной минус — "
     "существующие активные клиенты стали покупать МЕНЬШЕ: -₪522K на 94 связках клиент-департамент, "
     "частично компенсировано ростом других существующих (+₪295K) и новыми (+₪108K).", False),
    ("Контекст: команда сдаранов почти не занимается выкладкой ICE BDD — этот товар идёт через "
     "отдельный канал OneSales (водитель = агент = мерч в одном лице). Поэтому колонка 'שם סדרן' "
     "здесь отражает территорию/клиента, а не того, кто фактически продаёт, и падение справедливее "
     "смотреть со стороны канала OneSales/route-sales (частота визитов, доступность SKU в машине), "
     "а не как результат работы мерчендайзинга.", False),
    ("", False),
    ("INTER — лучший результат (+51.6%, +₪1.97M) — опыт стоит масштабировать", True),
    ("Рост держится на двух опорах одновременно: 42 новых клиента (+₪740K) И рост существующих "
     "(+₪1.33M на 98 связках), при почти нулевом оттоке (₪757). Это здоровый, широкий рост, "
     "не разовый эффект.", False),
    ("Топ-3 драйвера по сумме, ТОЛЬКО same-store (без новых клиентов, иначе цифра врёт): "
     "אוקסנה קיצ'ייב (+₪528K, +41.4%), יבגני בלקיטני (+₪301K, +62.3%), יבגני זמשה (+₪299K, +59.1%).", False),
    ("Рекомендация по масштабированию: разобрать конкретно, что делают אוקסנה קיצ'ייב, יבגני בלקיטני "
     "и יבגני זמשה в INTER (частота визитов, выкладка, работа с новыми точками) и по возможности "
     "перенести этот подход на сдаранов с более слабым ростом в тех же департаментах "
     "(FORMULA/ICE MISH, где эти же сдараны тоже работают напрямую).", False),
    ("", False),
    ("По сдаранам считаем ТОЛЬКО same-store (like-for-like), без ICE BDD (канал OneSales, "
     "сдараны его не ведут) — так принято в FMCG-дистрибуции", True),
    ("Если считать % роста сдарана вместе с новыми клиентами, картина искажается: "
     "יבגני זמשה по общей сумме (без ICE BDD) даёт +36.2%, но на своей РЕАЛЬНОЙ (существующей) "
     "базе same-store рост — +19.3%. Он всё равно хороший результат (2-й по величине same-store "
     "среди 6 сдаранов), но не такой выдающийся, как blended-цифра показывает: почти ₪627K его "
     "'роста' — это новые клиенты, а их сдарану НАЗНАЧАЮТ, он их не приводит сам.", False),
    ("Из-за этого меняется даже лидер: у אלכסיי נובוסלסקי вообще нет новых клиентов за период, "
     "поэтому его blended-показатель равен same-store — и оба выше, чем same-store יבגני זמשה: "
     "+21.1% против +19.3%. По 'чистой' работе с базой лидер — он, а не тот, у кого просто больше "
     "новых точек в назначенной территории.", False),
    ("Обратный случай ещё важнее: יגאל רייכמן показывал '+128.5%' по общей сумме — почти целиком "
     "это один новый клиент (₪64K). На своей реальной базе (5 связок, ₪49K) он в небольшом "
     "минусе: -1.9% same-store. Общая цифра полностью маскировала слабый результат на "
     "существующих клиентах.", False),
    ("Полная same-store таблица по всем сдаранам — лист 'По сдаранам', колонка "
     "'% изменение (Same-Store)'; вклад новых клиентов вынесен туда же отдельной колонкой, "
     "чтобы не путать два разных навыка — 'удержать и вырастить' и 'получить новую территорию'.", False),
]
rr = 3
for text, is_header in reco_text:
    if not text:
        rr += 1
        continue
    ws9.write(rr, 2, text, reco_header_fmt if is_header else reco_body_fmt)
    ws9.set_row(rr, 22 if is_header else 60)
    rr += 1

wb.close()
print("SAVED:", OUT)
print("total_ly", total_ly, "total_now", total_now, "pct", total_pct)
print("top dept growth:", by_dept[0]["key"], by_dept[0]["delta"])
print("top dept decline:", by_dept[-1]["key"], by_dept[-1]["delta"])
print("top sadran growth:", by_sadran[0]["key"], by_sadran[0]["delta"])
print("top sadran decline:", by_sadran[-1]["key"], by_sadran[-1]["delta"])
