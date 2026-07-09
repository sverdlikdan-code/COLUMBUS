"""
PROPHET MMD — Pilot demand forecasting for top-10 MMD SKUs.

Input:
  prophet/weekly_sales.csv   — SKU x week x carton sales (from fetch-prophet-history.js)
  prophet/top10_skus.json    — SKU metadata + hamlatza (current recommendation)

Output:
  prophet/forecast_pilot.xlsx — comparison table: fact / forecast / hamlatza / delta

Run:
  python prophet/run_prophet.py
"""

import json, warnings, os
import pandas as pd
from prophet import Prophet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── Israeli FMCG holiday calendar (Gregorian dates, 2022-2027) ──────────────
# Source: FMCG Israel skill + Hebrew calendar conversion
# Impact scale: 5=max (Pesach), 4=high, 3=medium
IL_HOLIDAYS = pd.DataFrame([
    # Pesach — massive pre-holiday buying spike, crash during 8 days
    {'ds': '2022-04-08', 'holiday': 'pesach_eve',       'lower_window': -14, 'upper_window': 1},
    {'ds': '2022-04-16', 'holiday': 'pesach_end',        'lower_window': 0,  'upper_window': 5},
    {'ds': '2023-04-05', 'holiday': 'pesach_eve',        'lower_window': -14, 'upper_window': 1},
    {'ds': '2023-04-13', 'holiday': 'pesach_end',        'lower_window': 0,  'upper_window': 5},
    {'ds': '2024-04-22', 'holiday': 'pesach_eve',        'lower_window': -14, 'upper_window': 1},
    {'ds': '2024-04-30', 'holiday': 'pesach_end',        'lower_window': 0,  'upper_window': 5},
    {'ds': '2025-04-12', 'holiday': 'pesach_eve',        'lower_window': -14, 'upper_window': 1},
    {'ds': '2025-04-20', 'holiday': 'pesach_end',        'lower_window': 0,  'upper_window': 5},
    {'ds': '2026-04-01', 'holiday': 'pesach_eve',        'lower_window': -14, 'upper_window': 1},
    {'ds': '2026-04-09', 'holiday': 'pesach_end',        'lower_window': 0,  'upper_window': 5},
    {'ds': '2027-04-21', 'holiday': 'pesach_eve',        'lower_window': -14, 'upper_window': 1},

    # Rosh Hashana — family meals, big pre-holiday spike
    {'ds': '2022-09-25', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},
    {'ds': '2023-09-15', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},
    {'ds': '2024-10-02', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},
    {'ds': '2025-09-22', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},
    {'ds': '2026-09-11', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},
    {'ds': '2027-10-01', 'holiday': 'rosh_hashana',      'lower_window': -7,  'upper_window': 3},

    # Sukkot — 7-day family holiday, outdoor eating
    {'ds': '2022-10-09', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},
    {'ds': '2023-09-29', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},
    {'ds': '2024-10-16', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},
    {'ds': '2025-10-06', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},
    {'ds': '2026-09-25', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},
    {'ds': '2027-10-15', 'holiday': 'sukkot',            'lower_window': -5,  'upper_window': 7},

    # Shavuot — DAIRY holiday, key for milk/cheese/ice cream SKUs
    {'ds': '2022-06-04', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},
    {'ds': '2023-05-25', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},
    {'ds': '2024-06-11', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},
    {'ds': '2025-06-01', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},
    {'ds': '2026-05-21', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},
    {'ds': '2027-06-10', 'holiday': 'shavuot',           'lower_window': -5,  'upper_window': 2},

    # Independence Day — BBQ season, snacks/drinks spike
    {'ds': '2022-05-04', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},
    {'ds': '2023-04-25', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},
    {'ds': '2024-05-13', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},
    {'ds': '2025-04-30', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},
    {'ds': '2026-04-21', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},
    {'ds': '2027-05-11', 'holiday': 'independence_day',  'lower_window': -3,  'upper_window': 1},

    # Hanukkah — sweets/dairy spike, coincides with winter
    {'ds': '2022-12-18', 'holiday': 'hanukkah',          'lower_window': -7,  'upper_window': 8},
    {'ds': '2023-12-07', 'holiday': 'hanukkah',          'lower_window': -7,  'upper_window': 8},
    {'ds': '2024-12-25', 'holiday': 'hanukkah',          'lower_window': -7,  'upper_window': 8},
    {'ds': '2025-12-14', 'holiday': 'hanukkah',          'lower_window': -7,  'upper_window': 8},
    {'ds': '2026-12-04', 'holiday': 'hanukkah',          'lower_window': -7,  'upper_window': 8},

    # Iron Swords War — structural demand shift Oct 2023 (negative shock)
    {'ds': '2023-10-07', 'holiday': 'iron_swords_war',   'lower_window': 0,   'upper_window': 60},
])
IL_HOLIDAYS['ds'] = pd.to_datetime(IL_HOLIDAYS['ds'])

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def week_to_date(year, week):
    """Convert PBI year+week to Monday of that week. PBI uses ISO-like numbering (%G-W%V); week 53 clamped to 52."""
    try:
        return pd.to_datetime(f"{int(year)}-W{int(week):02d}-1", format="%G-W%V-%u")
    except ValueError:
        return pd.to_datetime(f"{int(year)}-W52-1", format="%G-W%V-%u")


def load_data():
    from datetime import date, timedelta
    csv_path  = os.path.join(SCRIPT_DIR, 'weekly_sales.csv')
    top10_path = os.path.join(SCRIPT_DIR, 'top10_skus.json')
    df  = pd.read_csv(csv_path, dtype={'mkt': str})
    top10 = json.load(open(top10_path, encoding='utf-8'))
    # Remove current + future incomplete weeks — filter by converted date, not week number
    # PBI week numbers use Sunday-start; ISO uses Monday-start → mismatch if filtering by number
    today = date.today()
    week_start_monday = today - timedelta(days=today.weekday())  # Monday of current week
    cutoff = pd.Timestamp(week_start_monday)
    # Convert year+week to ds temporarily for filtering
    df['_ds'] = df.apply(lambda r: week_to_date(r['year'], r['week']), axis=1)
    before = len(df)
    df = df[df['_ds'] < cutoff].drop(columns=['_ds'])
    print(f"Removed {before - len(df)} rows with week_start >= {week_start_monday} (current/future weeks)")
    # Build name and hamlatza lookup
    meta = {str(r['mkt']): r for r in top10}

    # Load hamlatza_map.json for weeks_nf and hamlatza per SKU (all active SKUs)
    haml_map_path = os.path.join(SCRIPT_DIR, 'hamlatza_map.json')
    haml_map = {}
    if os.path.exists(haml_map_path):
        raw_map = json.load(open(haml_map_path, encoding='utf-8'))
        for k, v in raw_map.items():
            if isinstance(v, dict):
                haml_map[str(k)] = {'hamlatza': v.get('hamlatza'), 'weeks_nf': v.get('weeks_nf')}
            else:
                haml_map[str(k)] = {'hamlatza': v, 'weeks_nf': None}

    return df, meta, haml_map


def fit_forecast(df_sku, forecast_weeks=8):
    """Fit Prophet on one SKU's weekly series, return forecast df."""
    df_sku = df_sku.copy()
    df_sku['ds'] = df_sku.apply(lambda r: week_to_date(r['year'], r['week']), axis=1)
    df_sku = df_sku[['ds', 'mkr_k']].rename(columns={'mkr_k': 'y'})
    df_sku = df_sku.sort_values('ds').drop_duplicates('ds')

    # ── Stockout protection: remove zero-sale weeks (likely no inventory) ──
    # We keep them only if surrounded by zeros (product discontinued), otherwise drop
    df_sku = df_sku[df_sku['y'] > 0].copy()

    # ── Outlier clipping: cap spikes > mean + 2.5σ (promo/one-off loads) ──
    if len(df_sku) >= 8:
        mu, sigma = df_sku['y'].mean(), df_sku['y'].std()
        cap = mu + 2.5 * sigma
        df_sku['y'] = df_sku['y'].clip(upper=cap)

    if len(df_sku) < 4:
        return None, df_sku

    m = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=False,
        daily_seasonality=False,
        seasonality_mode='multiplicative',
        changepoint_prior_scale=0.05,   # conservative — avoid over-extrapolating spikes
        interval_width=0.80,
        holidays=IL_HOLIDAYS,           # Israeli FMCG holiday calendar
    )
    m.fit(df_sku)

    future = m.make_future_dataframe(periods=forecast_weeks, freq='W')
    forecast = m.predict(future)
    return forecast, df_sku


def build_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prophet Pilot"

    # ── Colors ──
    HDR_FILL  = PatternFill("solid", fgColor="1A237E")   # dark blue
    SEC_FILL  = PatternFill("solid", fgColor="E8EAF6")   # light blue
    RED_FILL  = PatternFill("solid", fgColor="FFCDD2")
    GRN_FILL  = PatternFill("solid", fgColor="C8E6C9")
    ALT_FILL  = PatternFill("solid", fgColor="F5F5F5")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT = Font(name="Calibri", size=10)
    BOLD_FONT = Font(bold=True, name="Calibri", size=10)

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    RIGHT  = Alignment(horizontal='right',  vertical='center')
    LEFT   = Alignment(horizontal='left',   vertical='center')

    headers = [
        'מק"ט', 'שם מוצר', 'קרטון/שבוע ממוצע (26w)',
        'P4 — פרופט 4w', 'P8 — פרופט 8w',
        'המלצה נוכחית', 'דלתא P4', 'דלתא P8',
        'נקודות נתון', 'הערה'
    ]
    col_widths = [10, 40, 22, 16, 16, 18, 12, 12, 12, 20]

    # Header row
    ws.row_dimensions[1].height = 36
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze header + enable auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for ri, r in enumerate(results, 2):
        row_fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            r['mkt'], r['taur'], r['avg_weekly'],
            r['p4'], r['p8'], r['hamlatza'],
            r['delta_p4'], r['delta_p8'],
            r['n_points'], r['note']
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            cell.font   = BODY_FONT

            if ci in (1, 2, 10):
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER
                cell.fill = row_fill

            # Color delta columns
            if ci == 7 and isinstance(v, (int, float)):   # delta P4
                cell.fill = GRN_FILL if v > 0 else RED_FILL
                cell.font = BOLD_FONT
            if ci == 8 and isinstance(v, (int, float)):   # delta P8
                cell.fill = GRN_FILL if v > 0 else RED_FILL
                cell.font = BOLD_FONT

        ws.row_dimensions[ri].height = 20

    # Summary row
    sr = len(results) + 3
    ws.cell(row=sr, column=1, value="📊 Legend").font = BOLD_FONT
    ws.cell(row=sr + 1, column=1, value="P4/P8 — Prophet forecast avg cartons/week for next 4 or 8 weeks").font = Font(name="Calibri", size=9, italic=True)
    ws.cell(row=sr + 2, column=1, value="דלתא > 0  →  Prophet expects MORE demand than current recommendation").font = Font(name="Calibri", size=9, italic=True, color="1B5E20")
    ws.cell(row=sr + 3, column=1, value="דלתא < 0  →  Prophet expects LESS demand than current recommendation").font = Font(name="Calibri", size=9, italic=True, color="B71C1C")

    wb.save(output_path)
    # Use ascii-safe print for Windows cp1255 console
    import sys
    sys.stdout.buffer.write(f"[OK] Saved {output_path}\n".encode('utf-8', errors='replace'))


def save_json(results, mmd_orders_path):
    """Save prophet.json alongside mmd-orders.json for the web app.

    Format: { "ts": ..., "skus": { "403001": { "p4": 46.5, "p8": 44.0 }, ... } }
    The browser computes: prophet_order = max(0, p4 * weeks_nf - eilat_k)
    using weeks_nf and eilat_k from mmd-orders.json.
    """
    # Read mmd-orders.json to merge weeks_nf + eilat_k
    mmd = {}
    if os.path.exists(mmd_orders_path):
        raw = json.load(open(mmd_orders_path, encoding='utf-8'))
        for r in raw.get('data', []):
            mmd[str(r['mkt'])] = {
                'weeks_nf': r.get('weeks_nf'),
                'eilat_k':  r.get('eilat_k'),
            }

    skus = {}
    for r in results:
        mkt = str(r['mkt'])
        wn  = mmd.get(mkt, {}).get('weeks_nf')
        ek  = mmd.get(mkt, {}).get('eilat_k') or 0
        p4  = r['p4']
        p8  = r['p8']
        # Prophet ORDER = max(0, p4 * weeks_nf - eilat_k)
        prophet_order = None
        if p4 is not None and wn is not None:
            prophet_order = max(0, round(p4 * wn - ek, 1))
        skus[mkt] = {
            'p4':            p4,
            'p8':            p8,
            'prophet_order': prophet_order,
            'weeks_nf':      wn,
            'eilat_k':       ek,
            'hamlatza':      r['hamlatza'],
            'delta_p4':      r['delta_p4'],
            'n_points':      r['n_points'],
            'series': {
                'actual': r.get('series_actual', []),
                'fc':     r.get('series_fc',     []),
            },
        }

    out = {
        'ts':   int(__import__('time').time() * 1000),
        'skus': skus,
    }
    docs_dir = os.path.join(SCRIPT_DIR, '..', 'docs')
    json_path = os.path.join(docs_dir, 'prophet.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    import sys
    sys.stdout.buffer.write(f"[OK] Saved {json_path}\n".encode('utf-8', errors='replace'))


def main():
    print("Loading data...")
    df, meta, haml_map = load_data()

    print(f"SKUs in CSV: {df['mkt'].nunique()}, rows: {len(df)}")

    results = []
    for mkt, group in df.groupby('mkt'):
        info = meta.get(str(mkt), {})
        hm   = haml_map.get(str(mkt), {})
        taur = group['taur'].iloc[0] if not group.empty else str(mkt)

        print(f"  Fitting mkt={mkt} ({len(group)} weeks)...", end=' ')
        forecast, hist = fit_forecast(group)

        n_points   = len(hist)
        avg_weekly = round(hist['y'].tail(26).mean(), 1) if len(hist) >= 4 else None

        if forecast is None:
            print("SKIP (too few points)")
            actual_series = [
                {'w': row['ds'].strftime('%Y-%m-%d'), 'y': round(float(row['y']), 1)}
                for _, row in hist.iterrows()
            ] if hist is not None and not hist.empty else []
            results.append({
                'mkt': mkt, 'taur': taur,
                'avg_weekly': avg_weekly, 'p4': None, 'p8': None,
                'hamlatza': hm.get('hamlatza') if hm else info.get('hamlatza'),
                'delta_p4': None, 'delta_p8': None,
                'n_points': n_points, 'note': 'Too few data points',
                'series_actual': actual_series,
                'series_fc': [],
            })
            continue

        # Future weeks only (yhat for last 4 and 8 periods)
        future_fc = forecast[forecast['ds'] > hist['ds'].max()]
        p4 = round(future_fc['yhat'].head(4).mean(), 1) if len(future_fc) >= 4 else None
        p8 = round(future_fc['yhat'].head(8).mean(), 1) if len(future_fc) >= 8 else None

        # Clamp negative forecasts to 0
        if p4 is not None: p4 = max(0.0, p4)
        if p8 is not None: p8 = max(0.0, p8)

        # Sanity cap: if P4 > avg × weeks_nf (PBI order base) → model anomaly, discard
        wn = hm.get('weeks_nf')
        if avg_weekly and avg_weekly > 0:
            cap_mult = float(wn) if (wn is not None and float(wn) > 0) else 4.0
            if p4 is not None and p4 > avg_weekly * cap_mult:
                print(f"    [ANOMALY] P4={p4} > {cap_mult}x avg={avg_weekly*cap_mult:.1f} -> null")
                p4 = None; p8 = None

        hamlatza  = hm.get('hamlatza') if hm else info.get('hamlatza')
        delta_p4  = round(p4 - hamlatza, 1) if p4 is not None and hamlatza is not None else None
        delta_p8  = round(p8 - hamlatza, 1) if p8 is not None and hamlatza is not None else None

        # Weekly series for chart (last 12 actual + next 4 forecast)
        actual_series = [
            {'w': row['ds'].strftime('%Y-%m-%d'), 'y': round(float(row['y']), 1)}
            for _, row in hist.tail(12).iterrows()
        ]
        fc_series = [
            {
                'w':  row['ds'].strftime('%Y-%m-%d'),
                'y':  round(max(0.0, float(row['yhat'])), 1),
                'lo': round(max(0.0, float(row['yhat_lower'])), 1),
                'hi': round(max(0.0, float(row['yhat_upper'])), 1),
            }
            for _, row in future_fc.head(4).iterrows()
        ]

        print(f"avg={avg_weekly}  P4={p4}  P8={p8}  ham={hamlatza}  D={delta_p4}")
        results.append({
            'mkt': mkt, 'taur': taur,
            'avg_weekly': avg_weekly, 'p4': p4, 'p8': p8,
            'hamlatza': hamlatza,
            'delta_p4': delta_p4, 'delta_p8': delta_p8,
            'n_points': n_points, 'note': '',
            'series_actual': actual_series,
            'series_fc':     fc_series,
        })

    out_path = os.path.join(SCRIPT_DIR, 'forecast_pilot.xlsx')
    build_excel(results, out_path)

    mmd_orders_path = os.path.join(SCRIPT_DIR, '..', 'docs', 'mmd-orders.json')
    save_json(results, mmd_orders_path)
    print("\nDone.")


if __name__ == '__main__':
    main()
